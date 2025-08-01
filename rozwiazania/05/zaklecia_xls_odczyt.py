from openpyxl import load_workbook

wb = load_workbook("zaklecia.xlsx")
arkusz = wb.active

for wiersz in arkusz.iter_rows(values_only=True):
    print(wiersz)